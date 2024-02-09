import logging
from pprint import pprint
from environs import Env
from typing import List, Dict, Any
import sqlite3    
import openpyxl
import datetime

from langchain_openai import ChatOpenAI
from langchain.memory import ConversationTokenBufferMemory
from langchain.prompts.prompt import PromptTemplate
from langchain.chains import ConversationChain


logging.basicConfig(
    filename="logging.txt",
    filemode="a",
    format="%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
    level=logging.DEBUG,
)

# import env config file
env = Env()
env.read_env()#'../.env', recurse=False)


class CustomConversationTokenBufferMemory(ConversationTokenBufferMemory):
    def clear(self):
        super().clear()
        buffer = self.chat_memory.messages
        curr_buffer_length = self.llm.get_num_tokens_from_messages(buffer)
        while curr_buffer_length > self.max_token_limit:
            buffer.pop(0)
            curr_buffer_length = self.llm.get_num_tokens_from_messages(buffer)
    
    extra_variables:List[str] = []

    @property
    def memory_variables(self) -> List[str]:
        """Will always return list of memory variables."""
        return [self.memory_key] + self.extra_variables

    def load_memory_variables(self, inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Return buffer with history and extra variables"""
        d = super().load_memory_variables(inputs)
        d.update({k:inputs.get(k) for k in self.extra_variables})        
        return d


async def update_info() -> tuple[str, str]:
    """ Обновление системного промпта и скрипта продаж

    Returns:
        tuple[str, str]: системный промпт, скрипт продаж
    """   

    excel_script_path = 'docs/sales_scripts.xlsx'
    wb = openpyxl.load_workbook(excel_script_path)
    ws = wb.active
    scripts_text= ''
    for row in ws.iter_rows(min_row=2):
        scripts_text += "Клиент:" + str(row[0].value) + '\n' + "AI ассистент:" + str(row[1].value) + '\n'
    scripts_text 
    
    # print('Считан скрипт продаж:\n', scripts_text)
    # with open('docs/train_script.txt', 'r') as f:
    #     scripts_text = f.read()
    
    with open('docs/system_prompt.txt', 'r') as f:
        system_prompt = f.read()
    #print('Считан системный промпт\n', system_prompt)
    
    return system_prompt, scripts_text


async def get_prompt() -> PromptTemplate:
    """Сборка шаблона промпта

    Returns:
        PromptTemplate: Шаблон промпта
    """        
    system_prompt, scripts_text = await update_info()
    template = f"""{system_prompt}\n
    Скрипт продаж:\n {scripts_text}\n"""
    template += """История общения с текущим клиентом: \n
    {history}\n
    Клиент: \n
    {input} \n
    AI ассистент:\n"""
    prompt_template = PromptTemplate(input_variables= ["history", "input"],
        template=template,)
    
    return prompt_template
        
# async def get_conversation_chain_history(prompt_template, verbose=False):

#     llm = ChatOpenAI(openai_api_key=env("OPENAI_TOKEN"), temperature=0.7)
#     # conversation = LLMChain(llm=llm, prompt=prompt_template)
#     conversation = ConversationChain(
#     llm=llm, verbose=verbose, prompt=prompt_template, memory=CustomConversationTokenBufferMemory(llm=llm, max_token_limit=2000, human_prefix="Клиент", ai_prefix="AI ассистент", extra_variables=["scripts_text", "system_prompt"]))
#     return conversation



# async def get_conversation_chain_window(prompt_template, verbose=False):
#     llm = ChatOpenAI(openai_api_key=env("OPENAI_TOKEN"), temperature=0.7)
#     conversation = ConversationChain(
#     llm=llm, verbose=verbose, prompt=prompt_template, memory=ConversationBufferWindowMemory(k=1, human_prefix="Клиент", ai_prefix="AI ассистент")
# )
#     return conversation


# async def get_conversation_chain_summary(prompt_template, verbose=False):
#     llm = ChatOpenAI(openai_api_key=env("OPENAI_TOKEN"), temperature=0.7)
#     conversation = ConversationChain(
#     llm=llm, verbose=verbose, prompt=prompt_template, memory=ConversationSummaryMemory(llm=llm)
# )
#     return conversation


async def get_user_conversation(user_id: int, verbose=False) -> ConversationChain:
    """Возвращает цепочку общения

    Args:
        user_id (int): телеграм ID пользователя
        verbose (bool, optional): Выводить в принт всю цепочку общения или нет. Defaults to False.

    Returns:
        ConversationChain: Цепочка общения
    """    
    prompt_template = await get_prompt()
    llm = ChatOpenAI(openai_api_key=env("OPENAI_TOKEN"), temperature=0.7)
    conversation = ConversationChain(
    llm=llm, verbose=verbose, prompt=prompt_template, memory=CustomConversationTokenBufferMemory(llm=llm, max_token_limit=2000, human_prefix="Клиент", ai_prefix="AI ассистент"))
    try:    
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()        
        # Извлекаем историю сообщений пользователя
        cursor.execute('SELECT input, output FROM (SELECT chat_id, input, output FROM chats WHERE user_id = ? and in_context = 1 ORDER BY chat_id DESC LIMIT 10) ORDER BY chat_id', (user_id, ))
        user_history= cursor.fetchall()
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
    finally:
        if connection:
            connection.close()
            logging.info(f"Выгрузка чата с пользователем {user_id}. Соединение с SQLite закрыто")
    if user_history:
        for input, output in user_history:
            conversation.memory.save_context({"input": input }, {"output": output}) # Заполнение цепочки сообщения контекстом общения с пользователем 
    return conversation
    
    
async def get_answer_from_llm(conversation: ConversationChain, input: str) -> str:
    """Получение ответов от OpenAI модели GPT

    Args:
        conversation (ConversationChain): Цепочка общения
        input (str): запрос пользователя

    Returns:
        str: ответ ИИ
    """
    try:    
        answer = await conversation.ainvoke(input=input)
        return answer['response']
    except:
        logging.error('Ошибка в получении ответа от ИИ')
        return None


async def insert_chat(user_id: int, input: str, output: str) -> bool:
    """Добавление текущего чата (вопрос-ответ) в БД

    Args:
        user_id (int): телеграм ID пользователя
        input (str): вопрос пользователя
        output (str): ответ ИИ ассистента

    Returns:
        bool: Статус выполнения функции
    """    
    try:
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()

        # Добавляем новое сообщение (вопрос-ответ) в базу
        cursor.execute('INSERT INTO chats (input, output, user_id, chat_datetime) VALUES (?, ?, ?, ?)', (input, output, user_id, datetime.datetime.now()))

        # Сохраняем изменения и закрываем соединение
        connection.commit()
        connection.close()
        status = True
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
        status = False
    finally:
        if connection:
            connection.close()
            logging.info("Чат с пользователем {user_id} обновлен в БД.Соединение с SQLite закрыто")
        return status
        

async def insert_user(user_id: int, username: str) -> bool:
    """Добавление нового пользователя в БД

    Args:
        user_id (int): телеграм ID пользователя
        username (str): username пользователя

    Returns:
        bool: статус выполнения функции
    """    
    try:
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()        
        # Проверяем наличие пользователя
        cursor.execute('SELECT user_id FROM users WHERE user_id = ?', (user_id, ))
        user_exists = cursor.fetchall()
        if user_exists == []:
        # Добавляем нового пользователя
            cursor.execute('INSERT INTO users (user_id, username) VALUES (?, ?)', (user_id, username))
        # Сохраняем изменения и закрываем соединение
            connection.commit()
        status = True
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
        status = False
    finally:
        if connection:
            connection.close()
            logging.info(f"Пользователь {user_id} добавлен в БД. Соединение с SQLite закрыто")
        return status
        
        
async def is_admin(user_id: int) -> bool:
    """Проверяет является ли пользователь администратором бота

    Args:
        user_id (int): телеграм ID пользователя 

    Returns:
        bool: Является ли пользователь админом
    """
    try:           
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()        
        # Проверяем наличие пользователя
        cursor.execute('SELECT user_id FROM users WHERE user_id = ? and isadmin = 1', (user_id, ))
        is_admin = cursor.fetchall()
        connection.commit()
        connection.close()
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
    finally:
        if connection:
            connection.close()
            logging.info(f"Проверка наличия пользователя {user_id} в админах. Соединение с SQLite закрыто")
    if is_admin == []:
        return False
    else:
        return True


async def make_admin(user_id: int) -> bool:
    """Включение пользователя в списки администраторов телеграм бота

    Args:
        user_id (int): телеграм ID пользователя 

    Returns:
        bool: Успешно добавлен
    """
    try:
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()        
        # Проверяем наличие пользователя
        cursor.execute('UPDATE users SET isadmin = 1 WHERE user_id = ?', (user_id, ))
        connection.commit()
        status = True
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
        status = False
    finally:
        if connection:
            connection.close()
            logging.info(f"Пользователь {user_id} включен в списки админов. Соединение с SQLite закрыто")
        return status
            

async def create_chats_report():
    """Формирует выгрузку с БД по чатам с клиентами для отправки админу
    """
    chat_report = None
    try:
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()        
        # Проверяем наличие пользователя
        cursor.execute('SELECT chat_datetime, chat_id, input, output, chats.user_id, username, isadmin FROM chats LEFT JOIN users ON chats.user_id = users.user_id;')
        chat_report = cursor.fetchall()
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
    finally:
        if connection:
            connection.close()
            logging.info(f"Отчет по чатам выгружен. Соединение с SQLite закрыто")
            import openpyxl
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.cell(row=1,column=1).value= 'datetime'
            sheet.cell(row=1,column=2).value = 'chat_id'
            sheet.cell(row=1,column=3).value = 'user_message'
            sheet.cell(row=1,column=4).value = 'ai_message'
            sheet.cell(row=1,column=5).value = 'user_id'
            sheet.cell(row=1,column=6).value = 'username'
            sheet.cell(row=1,column=7).value = 'is_admin'
            for enum, values in enumerate(chat_report):
                for col, value in enumerate(values):
                    sheet.cell(row=enum+2,column=col+1).value = value
            wb.save('docs/chat_history.xlsx')
            return chat_report
        
        
async def delete_self_context(user_id: int) -> bool:
    """
    Удаление контекста общения с ИИ у одного пользователя
    Args:
        user_id (int): телеграм ID пользователя 
    Returns:
        bool: Результат выполнения
    """
    try:
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()        
        # Проверяем наличие пользователя
        cursor.execute('UPDATE chats SET in_context = 0 WHERE user_id = ?', (user_id, ))
        connection.commit()
        state = True
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
        state = False
    finally:
        if connection:
            connection.close()
            logging.info(f"Контекст общения с пользователем {user_id} очищен. Соединение с SQLite закрыто")
        return state
            
            
async def delete_all_context() -> bool:
    """
    Удаление контекста общения с ИИ у одного пользователя

    Returns:
        bool: Результат выполнения
    """
    try:
        connection = sqlite3.connect('seller_bot.db')
        cursor = connection.cursor()        
        # Проверяем наличие пользователя
        cursor.execute('UPDATE chats SET in_context = 0')
        connection.commit()
        status = True
    except sqlite3.Error as error:
        logging.info("Ошибка при работе с SQLite", error)
        status = True
    finally:
        if connection:
            connection.close()
            logging.info(f"Контекст общения со всеми пользователями очищен. Соединение с SQLite закрыто")  
        return status     